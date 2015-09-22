#coding:utf-8

from openpyxl import load_workbook
from  openpyxl.cell  import  get_column_letter
from  openpyxl.writer.excel  import  ExcelWriter
from openpyxl.workbook import Workbook  

bid = [3716,3715,3714,3711,2279, 2874, 2891, 2686, 3433, 32, 1618, 2941, 2888, 2279,
       2224, 453, 2867, 2880, 2351, 2907, 2906, 2942, 3215, 3123, 2902, 2897, 2904, 50,
       11, 14, 2893, 23, 2341, 2891, 1619, 2865, 2890, 2901, 1673, 504, 2896, 108, 216,
       1, 455, 1238, 956, 943, 958, 1218, 533, 539, 542, 540, 538, 325, 607]


wb = load_workbook(filename='bookLIst.xlsx')
sheet_ranges = wb['Sheet1']
#
#print sheet_ranges['A2'].value
count_row = 1
count_col = 1
wb1 = Workbook()
ws1 = wb1.worksheets[0 ]
ew = ExcelWriter(workbook = wb1) 
ws = wb.active
row_count = 1
for row in ws.iter_rows("A1:AU3077"):
    try:
        value = int(row[0].value)
    except:
        pass
    for i in bid:
        if value == i:
            for i in row:
                col = get_column_letter(count_col)
                ws1.cell("%s%s" %(col, count_row)).value = i.value
                count_col += 1
            count_row += 1
            count_col = 1
dest_filename = r"result.xlsx"            
ew.save(filename = dest_filename)
